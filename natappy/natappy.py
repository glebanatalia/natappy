
import os
import time
import pickle
import json
import threading
import requests
import concurrent.futures
import textwrap
import xlsxwriter
import urllib
import openpyxl
import numpy as np
import pandas

from time import gmtime, strftime
from datetime import datetime
from datetime import timedelta
from os.path import expanduser
from functools import partial
from itertools import cycle
from contextlib import suppress
from PIL import Image

from selenium import webdriver
from selenium.common.exceptions import WebDriverException

import tweepy

import fbshares
import newpdf


home = expanduser("~")

APP_FOLDER = '{}/Desktop/natappy/bloggers.xlsx'.format(home)
INSTAGRAM_POSTS_LIMIT = 300
LOCK = threading.Lock()

facebook_reshares = {}
twitter_reshares = {}
links = []


def screenname_to_linkname(blogger, media):
    a = pandas.read_excel(APP_FOLDER)
    b = pandas.DataFrame(a)
    return list(b[b['Blogger'] == blogger][media])[0]


def linkname_to_screenname(blogger, media):
    a = pandas.read_excel(APP_FOLDER)
    b = pandas.DataFrame(a)
    return list(b[b[media] == blogger]['Blogger'])[0]


def username_to_id(blogger):
    a = pandas.read_excel(APP_FOLDER)
    b = pandas.DataFrame(a)
    return list(b[b['Instagram'] == blogger]['Instagram ID'])[0]


def request_more_media(next_, direction):
    more_media = requests.get(next_).json()
    if 'paging' in more_media and more_media['paging'] and direction in more_media['paging'] and more_media['paging'][direction]:
        more_data = more_media['data']
        more_next_ = more_media['paging'][direction]
        return more_data, more_next_
    if 'paging' in more_media and more_media['paging'] and not more_media['paging'][direction]:
        more_data = more_media['data']
        return more_data, None
    if 'paging' not in more_media and not more_media['data']:
        return [], None


def browse_posts_in_direction(root, direction):
    counter = 0
    collected = []
    while root and counter < 30:
        answer = request_more_media(root, direction=direction)
        collected.extend(answer[0])
        root = answer[1]
        counter += 1
    return collected


def pagination_fb(blogger, fromdate, todate, logg_details, dirname):
    if blogger == 'nothing':
        return [], 0
    until = "%s/%s/%s" % (todate[0], todate[1], todate[2])
    since = "%s/%s/%s" % (fromdate[0], fromdate[1], fromdate[2])
    link = "https://graph.facebook.com/v2.5/%s?fields=likes,posts.until(%s).since(%s){link,message,shares,status_type,created_time,likes.summary(true),comments.summary(true)}&access_token=%s" % (
        blogger, until, since, logg_details['token'])
    try:
        first_media = requests.get(link).json()
        next_, previous_, recent_media, followers = None, None, [], 0
        if 'posts' in first_media and first_media['posts']:
            followers = first_media['likes']
            recent_media = first_media['posts']['data']
            if 'paging' in first_media['posts']:
                next_ = first_media['posts']['paging'].get('next', None)
                previous_ = first_media['posts']['paging'].get('previous', None)
        else:
            print('blogger:', blogger, 'did not post anything between', since, 'and', until)
            link = "https://graph.facebook.com/v2.5/%s?fields=likes&access_token=%s" % (blogger, logg_details[
                                                                                        'token'])
            return [], requests.get(link).json()['likes']
        recent_media.extend(browse_posts_in_direction(root=next_, direction='next'))
        recent_media.extend(browse_posts_in_direction(root=previous_, direction='previous'))
        print('Facebook', len(recent_media), 'for', blogger, 'checked')
        return recent_media, followers
    except Exception as e:
        print('Exception while prcessing Facebook: {}'.format(e))
    return [], 0


def get_followers_count_ig(blogger, logg_details):
    try:
        url = 'http://instagram.com/' + blogger
        media = requests.get(url).text
        where = media.find('window._sharedData')
        half = (media[where + 21:])
        where2 = half.find(';</script>')
        content = half[:where2]
        to_process = (json.loads(content)['entry_data']['ProfilePage'])
        return to_process[0]['user']['followed_by']['count']
    except Exception as e:
        print('instagram followers count error:', e)
        return 0


def user_recent_media(username, max_id):
    url = 'http://instagram.com/' + username + '/media' + \
        ('?&max_id=' + max_id if max_id is not None else '')
    media = json.loads(requests.get(url).text)
    recent_media = media['items']
    if 'more_available' not in media or media['more_available'] is False:
        max_id = None
    else:
        max_id = media['items'][-1]['id']
    return recent_media, max_id


def pagination_ig(blogger, autho, fromdate, logg_details):
    if blogger == 'nothing':
        return None
    recent_media, next_ = user_recent_media(blogger, None)
    counter = 1
    from_ = datetime(fromdate[0], fromdate[1], fromdate[2])
    day_before_from = from_ - timedelta(1)
    try:
        while next_ and counter < INSTAGRAM_POSTS_LIMIT:
            more_media, next_ = user_recent_media(blogger, next_)
            times_ = [datetime.fromtimestamp(int(post['created_time'])) for post in more_media]
            if [time for time in times_ if time >= day_before_from]:
                recent_media.extend(more_media)
                counter += 1
            else:
                counter = 100000000000
        print('Instagram', len(recent_media), 'posts checked for blogger:', blogger)
        return recent_media
    except Exception as e:
        print(e)
        pass


def pagination_tw_inside(blogger, twapi, replies=None, retweets=None):
    recent_media = []
    followers = 0
    for tweet in tweepy.Cursor(twapi.user_timeline, twapi, screen_name=blogger, count=200, include_rts=True).items(limit=TWITTER_POSTS_LIMIT):
        followers = tweet.author.followers_count
        if hasattr(tweet, 'retweeted_status'):
            retweeted = True
            likes = tweet.retweeted_status.favorite_count
            retweet = tweet.retweeted_status.retweet_count
        else:
            retweeted = False
            likes = tweet.favorite_count
            retweet = tweet.retweet_count
        replied = True if tweet.in_reply_to_status_id_str != None and tweet.in_reply_to_screen_name != None else False
        tweet_link = 'https://twitter.com/%s/status/%s' % (blogger, tweet.id)
        tweet_tags = ['#' + hashtag['text'] for hashtag in tweet.entities.get(
            'hashtags')] + ['@' + mension['screen_name'] for mension in tweet.entities.get('user_mentions')]
        tweet_dict = {'tweet_id': tweet.id,
                      'text': str(tweet.text),
                      'retweeted': retweeted,
                      'likes': likes,
                      'retweet': retweet,
                      'created_at': str(tweet.created_at),
                      'replied': replied,
                      'in_reply_to_user_ID_str': str(tweet.in_reply_to_screen_name),
                      'in_reply_to_status_id': str(tweet.in_reply_to_status_id_str),
                      'link': str(tweet_link),
                      'tags': list(tweet_tags)}
        recent_media += [tweet_dict]
    print('Twitter', len(recent_media), 'posts checked for blogger:', blogger)
    return recent_media, followers


def pagination_tw(blogger, twapi, replies, retweets):
    if blogger == 'nothing':
        return [], 0
    counter = 0
    while counter < 45:
        try:
            return pagination_tw_inside(blogger, twapi, replies, retweets)
        except Exception as e:
            print('Can not connect host. Next trial in 30 sec. Exception:', e)
            time.sleep(30)
            return pagination_tw(blogger, twapi, replies, retweets)
            counter += 1
    return [], 0


def twitter_passwords_():
    passwords = []
    file = ([line.lstrip('\\fs48 ').rstrip('\\\n').split(',') for line in open(
        '%s/Desktop/natappy/passwords/twitter_passwords.txt' % home, 'r') if ',' in line])
    for line in file:
        appaswords = {'consumer_key': line[0],
                      'consumer_secret': line[1],
                      'token': line[2],
                      'token_secret': line[3],
                      'name': line[4],
                      'username': line[5],
                      'password': line[6]}
        passwords.append(appaswords)
    return passwords


twitter_passwords = twitter_passwords_()


def instagram_passwords_():
    passwords = []
    file = ([line.lstrip('\\fs48 ').rstrip('\\\n').split(',') for line in open(
        '%s/Desktop/natappy/passwords/instagram_passwords.txt' % home, 'r') if ',' in line])
    for line in file:
        appaswords = {}
        appaswords['client_id'] = line[0]
        appaswords['client_secret'] = line[1]
        appaswords['token'] = line[2]
        appaswords['name'] = line[3]
        appaswords['username'] = line[4]
        appaswords['password'] = line[5]
        passwords.append(appaswords)
    return passwords


instagram_passwords = instagram_passwords_()


def facebook_passwords_():
    passwords = []
    file = ([line.lstrip('\\fs48 ').rstrip('\\\n').split(',') for line in open(
        '%s/Desktop/natappy/passwords/facebook_passwords.txt' % home, 'r') if ',' in line])
    for line in file:
        appaswords = {}
        appaswords['expires'] = line[0]
        appaswords['token'] = line[1]
        appaswords['name'] = line[2]
        appaswords['username'] = line[3]
        appaswords['password'] = line[4]
        passwords.append(appaswords)
    return passwords


facebook_passwords = facebook_passwords_()


def logg():
    return {'Facebook': cycle(facebook_passwords), 'Instagram': cycle(instagram_passwords), 'Twitter': cycle(twitter_passwords)}


TEMP = logg()


def get_password(platform):
    with LOCK:
        password = next(TEMP[platform])
    return password


def process_blogger_facebook(blogger, fromdate, todate, logg_details, excludetags, dirname):
    print('Processing blogger on Facebook:', blogger)
    recent_media, followers = pagination_fb(blogger, fromdate, todate, logg_details, dirname)
    autho = None
    return recent_media, followers, autho


def process_blogger_instagram(blogger, fromdate, logg_details, excludetags):
    print('Processing blogger on Instagram:', blogger)
    autho = InstagramAPI(access_token=logg_details['token'], client_id=logg_details[
                         'client_id'], client_secret=logg_details['client_secret'])
    followers = get_followers_count_ig(blogger, logg_details=logg_details)
    recent_media = pagination_ig(blogger, autho=autho, fromdate=fromdate, logg_details=logg_details)
    return recent_media, followers, autho


def process_blogger_twitter(blogger, fromdate, logg_details, replies, retweets, excludetags):
    print('Processing blogger on Twitter:', blogger)
    auth = tweepy.AppAuthHandler(logg_details['consumer_key'], logg_details['consumer_secret'])
    autho = tweepy.API(auth, wait_on_rate_limit=True,
                       wait_on_rate_limit_notify=True)
    if (not autho):
        print("Can't Authenticate")

    if autho.rate_limit_status()['resources']['statuses']['/statuses/user_timeline']['remaining'] < 10:
        raise Exception(
            'You run out of rate limits. Application must close now. Please try again in 30 minutes.')
    recent_media, followers = pagination_tw(blogger, autho, replies, retweets)
    print('Twitter limits left:', autho.rate_limit_status()['resources'][
          'statuses']['/statuses/user_timeline']['remaining'])
    return recent_media, followers, autho


def process_blogger(blogger, platform, tags, repformat, fromdate, todate, followers_data, replies, retweets, excludetags, dirname):
    blogger = screenname_to_linkname(blogger, platform)
    logg_details = get_password(platform)
    if platform == 'Facebook':
        blogger_posts_info = process_blogger_facebook(
            blogger, fromdate, todate, logg_details=logg_details, excludetags=excludetags, dirname=dirname)
    if platform == 'Instagram':
        blogger_posts_info = process_blogger_instagram(
            blogger, fromdate, logg_details=logg_details, excludetags=excludetags)
    if platform == 'Twitter':
        blogger_posts_info = process_blogger_twitter(
            blogger, fromdate, logg_details=logg_details, replies=replies, retweets=retweets, excludetags=excludetags)
    recent_media = blogger_posts_info[0]
    followers = blogger_posts_info[1]
    try:
        followers_data[linkname_to_screenname(blogger, platform)][platform] += followers
    except:
        followers_data[blogger] = {'Facebook': 0, 'Instagram': 0, 'Twitter': 0}
        followers_data[blogger][platform] += followers
    autho = blogger_posts_info[2]
    if recent_media:
        with concurrent.futures.ThreadPoolExecutor(max_workers=20) as executor:
            result = list(executor.map(partial(process_post, tags=tags, autho=autho, platform=platform, repformat=repformat, blogger=blogger, followers=followers,
                                               fromdate=fromdate, todate=todate, replies=replies, retweets=retweets, excludetags=excludetags, dirname=dirname), recent_media))
 #   use if something broke
 #   result = list(map(partial(process_post, blogger=blogger , platform=platform, tags=tags, followers=followers), recent_media))
            result_filtered = {k: v for k, v in filter(lambda x: x is not None, result)}

            return blogger, result_filtered


def get_post_with_tags(bloggers, platform, tags, repformat, fromdate, todate, followers_data, replies, retweets, excludetags, dirname):
    executor_map = map
    if True:
        executor = concurrent.futures.ThreadPoolExecutor(max_workers=20)
        executor_map = executor.map
    results = executor_map(partial(process_blogger, platform=platform, tags=tags, repformat=repformat, fromdate=fromdate, todate=todate,
                                   followers_data=followers_data, replies=replies, retweets=retweets, excludetags=excludetags, dirname=dirname), bloggers)
    results_bloggers = {k: v for k, v in filter(lambda x: x is not None, results)}
    return results_bloggers


def process_post_facebook(post, blogger, tags, followers, from_, to_, excludetags, dirname):
    if 'message' in post and [tag for tag in tags if tag.lower() in post['message'].lower()] and not [excludetag for excludetag in excludetags if excludetag.lower() in post['message'].lower()]:
        print('Facebook posts matches criteria')

        try:
            likes = post['likes']['summary']['total_count']
        except:
            likes = 0
        try:
            shares = post['shares']['count']
        except:
            shares = 0
        try:
            comments = post['comments']['summary']['total_count']
        except:
            comments = 0
        if post['status_type'] not in ['shared_story', 'added_video', 'mobile_status_update']:
            try:
                link = post['link']
                note = post['status_type']
            except:
                link = 'https://www.facebook.com/%s' % post['id']
                note = post['status_type']
        if post['status_type'] == 'shared_story':
            link = 'https://www.facebook.com/%s' % post['id']
            if 'youtube' in post['link'] or 'youtu.be' in post['link']:
                note = 'shared_utube,%s' % str(len(textwrap.wrap(post['message'], 68)))
            else:
                note = 'shared_story'
        if post['status_type'] == 'added_video':
            #link='https://www.facebook.com/%s/videos' %blogger
            # note='added_video'
            link = post['link']
            note = 'shared_story'
        if post['status_type'] == 'mobile_status_update':
            note = 'mobile_status_update,%s' % str(len(textwrap.wrap(post['message'], 78)))
            link = 'https://www.facebook.com/%s' % post['id']

        take_screenshot(link, post['id'], note)
        if shares > 0:
            if not blogger in facebook_reshares:
                facebook_reshares[blogger] = {}
            if not post['id'] in facebook_reshares[blogger]:
                facebook_reshares[blogger][post['id']] = {}

        eng = 100 * ((likes + comments + shares) / followers)
        return post['id'], {'text': str(post['message']), 'shares': shares, 'likes': likes, 'created at': str(post['created_time'])[:16].replace('T', ' '), 'comments': comments, 'link': link, 'engagement': str(("%.2f" % eng)), 'followers': followers}


def process_post_instagram(post, tags, followers, from_, to_, excludetags):
    time_ = datetime.fromtimestamp(int(post['created_time']))
    try:
        #   print (post['caption'])
        if post['caption']['text'] and [tag for tag in tags if tag.lower() in post['caption']['text'].lower()] and to_ >= time_ >= from_ and not [excludetag for excludetag in excludetags if excludetag.lower() in post['caption']['text'].lower()]:
            print('Instagram post matches criteria')
            take_screenshot(str(post['link']), post['id'], note='regular')
            eng = 100 * (post['likes']['count'] + post['comments']['count']) / followers
            return post['id'], {'text': str(post['caption']['text']), 'created at': str(time_), 'likes': post['likes']['count'], 'comments': post['comments']['count'], 'link': post['link'], 'shares': 0, 'engagement': str(("%.2f" % eng)), 'followers': followers}
    except AttributeError:
        pass
    except Exception as e:
        print(post['caption'])
        print(e)
        pass


def process_post_twitter(post, blogger, tags, followers, from_, to_, replies, retweets, excludetags):
    time_ = datetime(int(str(post['created_at'])[:10].split(
        '-')[0]), int(str(post['created_at'])[:10].split('-')[1]), int(str(post['created_at'])[:10].split('-')[2]))
    if [tag for tag in tags if tag.lower() in post['text'].lower()] and to_ >= time_ >= from_ and not [excludetag for excludetag in excludetags if excludetag.lower() in post['text'].lower()]:
        tweet_link = 'https://twitter.com/%s/status/%s' % (blogger, str(post['tweet_id']))
        eng = 100 * (int(post['likes']) + int(post['retweet'])) / followers
        if replies and retweets:
            print('Twitter post matches criteria')
            if post['retweet'] > 0:
                if not blogger in facebook_reshares:
                    twitter_reshares[blogger] = {}
                if not post['tweet_id'] in twitter_reshares[blogger]:
                    twitter_reshares[blogger][post['tweet_id']] = {}
            take_screenshot(tweet_link, str(post['tweet_id']), note='regular')
            return int(post['tweet_id']), {'text': post['text'], 'likes': post['likes'], 'shares': post['retweet'], 'created at': str(post['created_at'])[:16], 'in_reply_to_user_ID_str': post['in_reply_to_user_ID_str'], 'in_reply_to_status_id': post['in_reply_to_status_id'], 'link': tweet_link, 'comments': 0, 'engagement': str(("%.2f" % eng)), 'followers': followers}
        if not replies and not retweets:
            if not post['retweeted'] and not post['replied']:
                if post['retweet'] > 0:
                    if not blogger in twitter_reshares:
                        twitter_reshares[blogger] = {}
                    if not post['tweet_id'] in twitter_reshares[blogger]:
                        print('One retweet collect:', post['tweet_id'])
                        twitter_reshares[blogger][post['tweet_id']] = {}
                print('Twitter post matches criteria')
                take_screenshot(tweet_link, str(post['tweet_id']), note='regular')
                return int(post['tweet_id']), {'text': post['text'], 'likes': post['likes'], 'shares': post['retweet'], 'created at': str(post['created_at'])[:16], 'in_reply_to_user_ID_str': post['in_reply_to_user_ID_str'], 'in_reply_to_status_id': post['in_reply_to_status_id'], 'link': tweet_link, 'comments': 0, 'engagement': str(("%.2f" % eng)), 'followers': followers}
        if not replies and retweets:
            if not post['replied']:
                if post['retweet'] > 0:
                    with LOCK:
                        if not blogger in twitter_reshares:
                            twitter_reshares[blogger] = {}
                        if not post['tweet_id'] in twitter_reshares[blogger]:
                            twitter_reshares[blogger][post['tweet_id']] = {}
                print('Twitter post matches criteria')
                take_screenshot(tweet_link, str(post['tweet_id']), note='regular')
                return int(post['tweet_id']), {'text': post['text'], 'likes': post['likes'], 'shares': post['retweet'], 'created at': str(post['created_at'])[:16], 'in_reply_to_user_ID_str': post['in_reply_to_user_ID_str'], 'in_reply_to_status_id': post['in_reply_to_status_id'], 'link': tweet_link, 'comments': 0, 'engagement': str(("%.2f" % eng)), 'followers': followers}
        if not retweets and replies:
            if not post['retweeted']:
                if post['retweet'] > 0:

                    if not blogger in twitter_reshares:
                        twitter_reshares[blogger] = {}
                    if not post['tweet_id'] in twitter_reshares[blogger]:
                        twitter_reshares[blogger][post['tweet_id']] = {}
                print('Twitter post matches criteria')
                take_screenshot(tweet_link, str(post['tweet_id']), note='regular')
                return int(post['tweet_id']), {'text': post['text'], 'likes': post['likes'], 'shares': post['retweet'], 'created at': str(post['created_at'])[:16], 'in_reply_to_user_ID_str': post['in_reply_to_user_ID_str'], 'in_reply_to_status_id': post['in_reply_to_status_id'], 'link': tweet_link, 'comments': 0, 'engagement': str(("%.2f" % eng)), 'followers': followers}
#         if not post['retweeted']:
#        if post['retweet']>0:
#
#            if not blogger in twitter_reshares:
#                twitter_reshares[blogger]={}
#            if not post['tweet_id'] in twitter_reshares[blogger]:
#                twitter_reshares[blogger][post['tweet_id']]={}
#        print ('twitter post matches criteria')
#        take_screenshot(tweet_link,str(post['tweet_id']),note='regular')
#        return int(post['tweet_id']), {'text':post['text'],'likes': post['likes'],'shares': post['retweet'],'created at':str(post['created_at'])[:16],'in_reply_to_user_ID_str': post['in_reply_to_user_ID_str'],'in_reply_to_status_id': post['in_reply_to_status_id'],'link':tweet_link,'comments':0,'engagement': str(("%.2f" %eng)),'followers': followers }
#


def process_post(post, blogger, platform, autho, tags, followers, repformat, fromdate, todate, replies, retweets, excludetags, dirname):
    from_ = datetime(fromdate[0], fromdate[1], fromdate[2])
    to_ = datetime(todate[0], todate[1], todate[2])
    if platform == 'Facebook':
        return process_post_facebook(post, blogger, tags, followers, from_=from_, to_=to_, excludetags=excludetags, dirname=dirname)
    if platform == 'Instagram':
        return process_post_instagram(post, tags, followers, from_=from_, to_=to_, excludetags=excludetags)
    if platform == 'Twitter':
        return process_post_twitter(post, blogger, tags, followers, from_, to_, replies, retweets, excludetags=excludetags)


def page_is_loaded(driver):
    return driver.find_element_by_tag_name("body") != None


def take_screenshot(url, image_name, note):
    links.append([str(url).replace("\'", '"'), str(image_name), str(note)])


def cut_screenshot_facebook(screenshot, note, platform, url):
    print('note:', note)
    time.sleep(3)
    try:
        print('Facebook screenshot memory size:', os.path.getsize(screenshot))
        if os.path.getsize(screenshot) > 200:
            img = Image.open(screenshot)
            print('Facebook image size:', img.size)
            if note == 'added_photos':
                img.crop((20, 20, img.size[0] - 20, img.size[1] - 20)).save('%s' % screenshot)
            if note == 'shared_story':
                if img.size[1] <= 734:
                    img.crop((190, 136, img.size[0] - 300,
                              img.size[1] - 170)).save('%s' % screenshot)
                    time.sleep(2)
                if img.size[1] > 734:
                    img.crop((190, 132, img.size[0] - 300,
                              img.size[1] - 100)).save('%s' % screenshot)
                    time.sleep(2)
            if note.startswith('mobile_status_update'):
                img.crop((190, 140, img.size[0] - 303, img.size[1] - 465 +
                          16 * int(note.split(',')[1]))).save('%s' % screenshot)
            if note == 'added_video':
                img.crop((25, 543, img.size[0] - 143, img.size[1] - 33)).save('%s' % screenshot)
                time.sleep(2)
            if note.startswith('shared_utube'):
                img.crop((190, 138, img.size[0] - 300, img.size[1] - 293 +
                          16 * int(note.split(',')[1]))).save('%s' % screenshot)
                time.sleep(2)
            if note != 'shared_story' and note != 'added_photos' and note != 'mobile_status_update' and note != 'added_video' and not note.startswith('shared_utube') and not note.startswith('mobile_status_update'):
                img.crop((0, 220, img.size[0] - 260, img.size[1] - 60)).save('%s' % screenshot)
                time.sleep(2)
        else:
            name = screenshot.split('/')[6]
            dirname = screenshot.split('/')[4]
            print('Broken Facebook screenshot. Going to take a new one')
            phantom(platform, url, name, note, dirname)

    except:
        name = screenshot.split('/')[6]
        dirname = screenshot.split('/')[4]
        print('Empty Facebook screenshot. Going to take a new one')
        phantom(platform, url, name, note, dirname)


def cut_screenshot_twitter(screenshot, note, url):
    img = Image.open(screenshot)
    print('Twitter image size:', img.size)
    img.crop((note[0], note[1], note[0] + note[2], note[1] + note[3])).save(screenshot)


def cut_screenshot_instagram(screenshot, note):
    img = Image.open(screenshot)
    print('Instagram image size:', img.size)
    img.crop((0, 77, img.size[0], img.size[1] - 120)).save("%s" % screenshot)
    time.sleep(2)


def cut_screenshot(screenshot, note, platform, url):
    try:
        if platform == 'Instagram':
            cut_screenshot_instagram(screenshot, note)
        elif platform == 'Twitter':
            cut_screenshot_twitter(screenshot, note, url)
        elif platform == 'Facebook':
            time.sleep(3)
            cut_screenshot_facebook(screenshot, note, platform, url)
    except IndexError:
        print('Uppss.. Something was wrong with screenshot. I should better grab another one.')
        name = screenshot.split('/')[6]
        dirname = screenshot.split('/')[4]
        phantom(platform, url, name, note, dirname)
    except Exception as e:
        print('Facebook screenshot to cut error:', e)
        pass


def phantom(platform, url, name, note, dirname):
    try:
        browser = webdriver.PhantomJS('/Users/crowdscores/desktop/phantomjs')
        if platform == 'Twitter':
            browser.get(url)
            time.sleep(6)
            browser.set_window_size(1366, 1500)
            try:
              #  element=browser.find_element_by_css_selector('div.expansion-container.js-expansion-container')
                element = browser.find_element_by_css_selector(
                    'div.permalink-inner.permalink-tweet-container')
                browser.save_screenshot('%s/Desktop/%s/%s_screenshots/%s.png' %
                                        (home, dirname, platform, name))
                size = element.size
                location = element.location
                note = [int(location['x']), int(location['y']),
                        int(size['width']), int(size['height'])]
                browser.quit()
              #  note=[50,50,1366,1500]
                cut_screenshot('%s/Desktop/%s/%s_screenshots/%s.png' %
                               (home, dirname, platform, name), url=url, note=note, platform=platform)

            except:

                print('Did not find selector.Will try once again, url:', url)
                browser.quit()
                phantom(platform, url, name, note, dirname)

        else:
            print(url)
            browser.get('%s' % url)
            time.sleep(15)
            browser.save_screenshot('%s/Desktop/%s/%s_screenshots/%s.png' %
                                    (home, dirname, platform, name))
            time.sleep(4)
            browser.quit()
            cut_screenshot('%s/Desktop/%s/%s_screenshots/%s.png' %
                           (home, dirname, platform, name), url=url, note=note, platform=platform)
    except WebDriverException:
        print('Connection problems... Will try once again')
        time.sleep(5)
        phantom(platform, url, name, note, dirname)
    except urllib.error.URLError:
        print('urllib error ignored. URL: ', url)
        pass


def collect_screenshot(url_name, dirname):
    url = url_name[0]
    name = url_name[1]
    note = url_name[2]
    if 'facebook' in url:
        phantom('Facebook', url, name, note, dirname)
    if 'instagram' in url:
        phantom('Instagram', url, name, note, dirname)
    if 'twitter' in url:
        phantom('Twitter', url, name, note, dirname)


def collect_screenshots(dirname):
    print('Collecting screenshots. Screenshots to collect:', len(links))
    tic = time.time()
    with concurrent.futures.ProcessPoolExecutor(max_workers=10) as executor:
        int in executor.map(partial(collect_screenshot, dirname=dirname), links)
    try:
        print('Screenshots collected. Elapsed:', time.time() - tic,
              'Average:', int(time.time() - tic) / len(links), 'per post')
    except:
        pass


def count_media_impression_blogger(all_media, followers_data, media, blogger):
    return followers_data[linkname_to_screenname(blogger, media)][media] * len(all_media[media][blogger])


def split_text(text):
    return textwrap.wrap(text, 60)


def excel_format(all_media, followers_data):
    post_in_excel_format = []
    for media in all_media.keys():
        for blogger in all_media[media].keys():
            counter = 1
            for post in all_media[media][blogger].keys():
                post_details = [media, linkname_to_screenname(blogger, media), followers_data[linkname_to_screenname(blogger, media)][media], str(post), all_media[media][blogger][post]['link'], all_media[media][blogger][post]['created at'][:10], all_media[media][blogger][post]['created at'][10:], all_media[media][blogger][post]['text'], all_media[
                    media][blogger][post]['likes'], all_media[media][blogger][post]['shares'], all_media[media][blogger][post]['comments'], all_media[media][blogger][post]['likes'] + all_media[media][blogger][post]['shares'] + all_media[media][blogger][post]['comments'], float(all_media[media][blogger][post]['engagement']), counter]
                post_in_excel_format.append(post_details)
                counter += 1
    return post_in_excel_format


def add_blogger_to_excel(blogger):
    wb = openpyxl.load_workbook('%s/Desktop/natappy/bloggers.xlsx' % home)
    sheet = wb.get_sheet_by_name('bloggers')
    position = sheet.get_highest_row()
    blogger = blogger.split(',')
    counter = 0
    for column in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        print(blogger)
        sheet[str(column) + str(position + 1)] = blogger[counter]
        print(sheet[str(column) + str(position + 1)])
        counter += 1
    wb.save('%s/Desktop/natappy/bloggers.xlsx' % home)


def excel(all_media, dirname, tags, followers_data):
    print('creating excel')

    workbook = xlsxwriter.Workbook('%s/Desktop/%s/Report.xlsx' % (home, dirname))
    worksheet = workbook.add_worksheet()
    fieldsnames = ('MEDIA', 'BLOGGER', 'FOLLOWERS', 'POSTS_ID', 'POSTS_LINK', 'CREATED_DATE', 'CREATED_TIME',
                   'POSTS_TEXT', 'POSTS_LIKES', 'POSTS_SHARES', 'POSTS_COMMENTS', 'POST_INTERACTIONS', 'POSTS_ENGAGEMENT')
#    columns=['A','B','C','D','E','F','G','H','I','J','K','L','M']
    col = 0
    for name in fieldsnames:
        #  sheet[str(columns[0])+str(row)]=name
        worksheet.write(0, col, name)
        col += 1
    posts = excel_format(all_media, followers_data)
    row = 1

    for post in posts:
        col = 0
        for name in post[:13]:
            new_image_name = ''.join([name_part for name_part in post[1]]
                                     ) + '_%s_(%s)' % (tags[0], post[13])
            with suppress(Exception):
                os.rename('%s/Desktop/%s/%s_screenshots/%s.png' % (home, dirname, post[0], post[
                          3]), '%s/Desktop/%s/%s_screenshots/%s.png' % (home, dirname, post[0], new_image_name))
            worksheet.write(row, col, name)
        #    sheet[str(columns[col])+str(row)]=name
            col += 1
        row += 1
    # wb.save('%s/Desktop/%s/Report.xlsx' %(home,dirname))
    workbook.close()

platshort = {'Facebook': 'FB', 'Instagram': 'IG', 'Twitter': 'TW', 'Blogs': 'BL'}


def addtional_Excel(dirname):
    wb = openpyxl.load_workbook('%s/Desktop/%s/Report.xlsx' % (home, dirname))
    df = pandas.read_excel('%s/Desktop/%s/Report.xlsx' % (home, dirname))
    fb = df[df["MEDIA"] == 'Facebook']
    ig = df[df["MEDIA"] == 'Instagram']
    tw = df[df["MEDIA"] == 'Twitter']
    TotalMedia = pandas.pivot_table(df, index=["MEDIA"], values=["POSTS_ID", "FOLLOWERS", "POST_INTERACTIONS", "POSTS_LIKES", "POSTS_COMMENTS", "POSTS_SHARES", "POSTS_ENGAGEMENT"], aggfunc={
                                    "POSTS_ID": len, "FOLLOWERS": np.sum, "POST_INTERACTIONS": np.sum, "POSTS_LIKES": np.sum, "POSTS_COMMENTS": np.sum, "POSTS_SHARES": np.sum, "POSTS_ENGAGEMENT": np.average})
    DayByDayTotal = pandas.pivot_table(df, index=["CREATED_DATE"], values=["POSTS_ID", "FOLLOWERS", "POST_INTERACTIONS", "POSTS_LIKES", "POSTS_COMMENTS", "POSTS_SHARES", "POSTS_ENGAGEMENT"], aggfunc={
                                       "POSTS_ID": len, "FOLLOWERS": np.sum, "POST_INTERACTIONS": np.sum, "POSTS_LIKES": np.sum, "POSTS_COMMENTS": np.sum, "POSTS_SHARES": np.sum, "POSTS_ENGAGEMENT": np.average})
    DayByDay = pandas.pivot_table(df, index=["MEDIA", "CREATED_DATE"], values=["POSTS_ID", "FOLLOWERS", "POST_INTERACTIONS", "POSTS_LIKES", "POSTS_COMMENTS", "POSTS_SHARES", "POSTS_ENGAGEMENT"], aggfunc={
                                  "POSTS_ID": len, "FOLLOWERS": np.sum, "POST_INTERACTIONS": np.sum, "POSTS_LIKES": np.sum, "POSTS_COMMENTS": np.sum, "POSTS_SHARES": np.sum, "POSTS_ENGAGEMENT": np.average})
    BloggersTotals = pandas.pivot_table(df, index=["BLOGGER"], values=["POSTS_ID", "FOLLOWERS", "POST_INTERACTIONS", "POSTS_LIKES", "POSTS_COMMENTS", "POSTS_SHARES", "POSTS_ENGAGEMENT"], aggfunc={
                                        "POSTS_ID": len, "FOLLOWERS": np.sum, "POST_INTERACTIONS": np.sum, "POSTS_LIKES": np.sum, "POSTS_COMMENTS": np.sum, "POSTS_SHARES": np.sum, "POSTS_ENGAGEMENT": np.average})
    Bloggers = pandas.pivot_table(df, index=["MEDIA", "BLOGGER"], values=["POSTS_ID", "FOLLOWERS", "POST_INTERACTIONS", "POSTS_LIKES", "POSTS_COMMENTS", "POSTS_SHARES", "POSTS_ENGAGEMENT"], aggfunc={
                                  "POSTS_ID": len, "FOLLOWERS": np.sum, "POST_INTERACTIONS": np.sum, "POSTS_LIKES": np.sum, "POSTS_COMMENTS": np.sum, "POSTS_SHARES": np.sum, "POSTS_ENGAGEMENT": np.average})
    HourByHourTotal = pandas.pivot_table(df, index=["CREATED_TIME"], values=["POSTS_ID", "FOLLOWERS", "POST_INTERACTIONS", "POSTS_LIKES", "POSTS_COMMENTS", "POSTS_SHARES", "POSTS_ENGAGEMENT"], aggfunc={
                                         "POSTS_ID": len, "FOLLOWERS": np.sum, "POST_INTERACTIONS": np.sum, "POSTS_LIKES": np.sum, "POSTS_COMMENTS": np.sum, "POSTS_SHARES": np.sum, "POSTS_ENGAGEMENT": np.average})
    HourByHour = pandas.pivot_table(df, index=["MEDIA", "CREATED_TIME"], values=["POSTS_ID", "FOLLOWERS", "POST_INTERACTIONS", "POSTS_LIKES", "POSTS_COMMENTS", "POSTS_SHARES", "POSTS_ENGAGEMENT"], aggfunc={
                                    "POSTS_ID": len, "FOLLOWERS": np.sum, "POST_INTERACTIONS": np.sum, "POSTS_LIKES": np.sum, "POSTS_COMMENTS": np.sum, "POSTS_SHARES": np.sum, "POSTS_ENGAGEMENT": np.average})
    MediaByBlogger = pandas.pivot_table(df, index=["MEDIA", "BLOGGER"], values=["POSTS_ID", "FOLLOWERS", "POST_INTERACTIONS", "POSTS_LIKES", "POSTS_COMMENTS", "POSTS_SHARES", "POSTS_ENGAGEMENT"], aggfunc={
                                        "POSTS_ID": len, "FOLLOWERS": np.sum, "POST_INTERACTIONS": np.sum, "POSTS_LIKES": np.sum, "POSTS_COMMENTS": np.sum, "POSTS_SHARES": np.sum, "POSTS_ENGAGEMENT": np.average})

    Facebook = pandas.pivot_table(fb, index=["BLOGGER"], values=["POSTS_ID", "FOLLOWERS", "POST_INTERACTIONS", "POSTS_LIKES", "POSTS_COMMENTS", "POSTS_SHARES", "POSTS_ENGAGEMENT"], aggfunc={
                                  "POSTS_ID": len, "FOLLOWERS": np.sum, "POST_INTERACTIONS": np.sum, "POSTS_LIKES": np.sum, "POSTS_COMMENTS": np.sum, "POSTS_SHARES": np.sum, "POSTS_ENGAGEMENT": np.average})
    Instagram = pandas.pivot_table(ig, index=["BLOGGER"], values=["POSTS_ID", "FOLLOWERS", "POST_INTERACTIONS", "POSTS_LIKES", "POSTS_COMMENTS", "POSTS_SHARES", "POSTS_ENGAGEMENT"], aggfunc={
                                   "POSTS_ID": len, "FOLLOWERS": np.sum, "POST_INTERACTIONS": np.sum, "POSTS_LIKES": np.sum, "POSTS_COMMENTS": np.sum, "POSTS_SHARES": np.sum, "POSTS_ENGAGEMENT": np.average})
    Twitter = pandas.pivot_table(tw, index=["BLOGGER"], values=["POSTS_ID", "FOLLOWERS", "POST_INTERACTIONS", "POSTS_LIKES", "POSTS_COMMENTS", "POSTS_SHARES", "POSTS_ENGAGEMENT"], aggfunc={
                                 "POSTS_ID": len, "FOLLOWERS": np.sum, "POST_INTERACTIONS": np.sum, "POSTS_LIKES": np.sum, "POSTS_COMMENTS": np.sum, "POSTS_SHARES": np.sum, "POSTS_ENGAGEMENT": np.average})
    writer = pandas.ExcelWriter('%s/Desktop/%s/ReportAnalyse.xlsx' % (home, dirname))
    writer.wb = wb
    writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
    TotalMedia.to_excel(writer, sheet_name='TotalMedia')
    DayByDayTotal.to_excel(writer, sheet_name='DayByDayTotal')
    DayByDay.to_excel(writer, sheet_name='DayByDay')
    BloggersTotals.to_excel(writer, sheet_name='BloggersTotal')
    Bloggers.to_excel(writer, sheet_name='Bloggers')
    HourByHourTotal.to_excel(writer, sheet_name='HourByHourTotal')
    HourByHour.to_excel(writer, sheet_name='HourByHour')
    Facebook.to_excel(writer, sheet_name='Facebook')
    Instagram.to_excel(writer, sheet_name='Instagram')
    Twitter.to_excel(writer, sheet_name='Twitter')
    MediaByBlogger.to_excel(writer, sheet_name='MediaByBlogger')
    writer.save()


def wait_for_file(name):
    counter = 0
    while True:
        if os.path.exists(name):
            return
        print('Waiting for file:', name)
        counter += 1
        if counter > 30:
            raise Exception('File could not be found for 30 seconds. File:{name}'.format(name=name))
        time.sleep(1)


def count_media_impression(all_media, media, followers_data):
    media_impression = 0
    for blogger in all_media[media]:
        media_impression += followers_data[linkname_to_screenname(
            blogger, media)][media] * len(all_media[media][blogger])
    return media_impression


def collect_facebook_reshares(dirname):
    if facebook_reshares:
        with suppress(Exception):
            os.makedirs('%s/Desktop/%s/Facebook_shares' % (home, dirname))
            print('Going to collect reposts results....')
            for blogger in facebook_reshares.keys():
                for post in facebook_reshares[blogger].keys():
                    only_post_id = post.split('_')[1]
                    post_reshares = fbshares.collecting_reshares_everything(only_post_id, dirname)
                    facebook_reshares[blogger][post] = post_reshares
    return facebook_reshares


def collect_twitter_reshares(dirname):
    if twitter_reshares:
        return fbshares.collect_retweets(twitter_reshares)


#==============================================================================
# main
#==============================================================================


def create_followers_data(screenbloggers, platforms):
    followers_data = {}
    for blogger in screenbloggers:
        followers_data[blogger] = {}
        for platform in platforms:
            followers_data[blogger][platform] = 0
    return followers_data


def cut_blog_screenshots(dirname, screenshot, blogger, counter):
    n = 4  # TODO depends on image size
    img = Image.open(screenshot)
    if img.size[1] > 8000:
        n = 5
    if img.size[1] > 13000:
        n = 8
    new_height = int(img.size[1] / n)
    print('Blog screenshot size:', img.size)
    pieces = []
    for i in range(n):
        img.crop((0, 0 + i * new_height, img.size[0], 0 + i * new_height + new_height)).save(
            '%s/Desktop/%s/Blogs_screenshots/%s_(%s)_(%s).png' % (home, dirname, blogger, counter, i))
        pieces.append('%s/Desktop/%s/Blogs_screenshots/%s_(%s)_(%s).png' %
                      (home, dirname, blogger, counter, i))
    return pieces


def grab_post_screenshot(dirname, link, blogger, counter):
    browser = webdriver.PhantomJS('/Users/crowdscores/desktop/phantomjs')
    if not link.startswith('http://'):
        link = 'http://' + link
    browser.get(link)
    time.sleep(10)
    lenOfPage = browser.execute_script(
        "window.scrollTo(0, document.body.scrollHeight);var lenOfPage=document.body.scrollHeight;return lenOfPage;")
    match = False
    print('Blog post screenshot in progress. URL:', link)
    while(match == False):
        lastCount = lenOfPage
        time.sleep(3)
        lenOfPage = browser.execute_script(
            "window.scrollTo(0, document.body.scrollHeight);var lenOfPage=document.body.scrollHeight;return lenOfPage;")
        if lastCount == lenOfPage:
            match = True
    time.sleep(3)
    browser.save_screenshot('%s/Desktop/%s/Blogs_screenshots/%s_(%s).png' %
                            (home, dirname, blogger, counter))
    browser.quit()
    n = cut_blog_screenshots(dirname, '%s/Desktop/%s/Blogs_screenshots/%s_(%s).png' %
                             (home, dirname, blogger, counter), blogger, counter)
    return n


def collect_blog_screenshots(dirname, blogposts):
    post_info = {}
    for blogger in blogposts:
        post_info[blogger] = {}
        counter = 0
        for link in blogposts[blogger]:
            post_info[blogger][link] = {}
            post_info[blogger][link]['number'] = counter
            post_info[blogger][link]['pieces'] = grab_post_screenshot(
                dirname, link, blogger, counter)
            counter += 1
    return post_info


def create_report(screenbloggers, platforms, tags, repformat, fromdate, todate, replies, retweets, excludetags, brandname, campaign, blogposts):
  #  print ('repformat:', repformat)

    all_media = {}
    post_info = {}

    platshorts = '_'.join([platshort[platform] for platform in platforms])
    dirname = '%s-%s-%s' % (platshorts, '_'.join([tag for tag in tags]),
                            str(strftime("%Y_%m_%d_%H_%M", gmtime())))
    pickle.dump(dirname, open(
        '{home}/Desktop/natappy/pickle/dirname.pickle'.format(home=home), 'wb'))
    with suppress(Exception):
        os.makedirs('%s/Desktop/%s' % (home, dirname))
    if repformat == 'Full':
        for platform in platforms:
            with suppress(Exception):
                os.makedirs('%s/Desktop/%s/%s_screenshots' % (home, dirname, platform))
    followers_data = create_followers_data(screenbloggers, platforms)

    for platform in [plat for plat in platforms if plat not in['Blogs']]:
        all_media[platform] = get_post_with_tags(
            screenbloggers, platform, tags, repformat, fromdate, todate, followers_data, replies, retweets, excludetags, dirname)
    pickle.dump(all_media, open(
        '{home}/Desktop/natappy/pickle/all_media.pickle'.format(home=home), 'wb'))
    pickle.dump(followers_data, open(
        '{home}/Desktop/natappy/pickle/followers.pickle'.format(home=home), 'wb'))
    pickle.dump(links, open('{home}/Desktop/natappy/pickle/links.pickle'.format(home=home), 'wb'))
    dirname = 'report'
    dirname = pickle.load(
        open('{home}/Desktop/natappy/pickle/dirname.pickle'.format(home=home), 'rb'))
    all_media = pickle.load(
        open('{home}/Desktop/natappy/pickle/all_media.pickle'.format(home=home), 'rb'))
    followers_data = pickle.load(
        open('{home}/Desktop/natappy/pickle/followers.pickle'.format(home=home), 'rb'))

    reshares = {'Facebook': {}, 'Twitter': {}}
    reshares['Facebook'] = collect_facebook_reshares(dirname)
    reshares['Twitter'] = collect_twitter_reshares(dirname)

    pickle.dump(reshares, open(
        '{home}/Desktop/natappy/pickle/reshares.pickle'.format(home=home), 'wb'))
    reshares = pickle.load(
        open('{home}/Desktop/natappy/pickle/reshares.pickle'.format(home=home), 'rb'))
    post_info = pickle.load(
        open('{home}/Desktop/natappy/pickle/postsinfo.pickle'.format(home=home), 'rb'))
    if reshares:
        fbshares.nice_reshares_excel(reshares, dirname)
    if repformat == 'Full':
        if 'Blogs' in platforms:
            print('Blogs')
            post_info = collect_blog_screenshots(dirname, blogposts)
            pickle.dump(post_info, open(
                '{home}/Desktop/natappy/pickle/postsinfo.pickle'.format(home=home), 'wb'))
        collect_screenshots(dirname)
    newpdf.pdf(all_media, tags, screenbloggers, fromdate, todate, dirname=dirname, platforms=platforms, followers_data=followers_data,
               size='full', brandname=brandname, campaign=campaign, reshares=reshares, blogposts=blogposts, post_info=post_info)
    excel(all_media, dirname, tags, followers_data)
    addtional_Excel(dirname)
    if repformat == 'Quick':
        if 'Blogs' in platforms:

            post_info = collect_blog_screenshots(dirname, blogposts)
            pickle.dump(post_info, open(
                '{home}/Desktop/natappy/pickle/postsinfo.pickle'.format(home=home), 'wb'))
        newpdf.pdf(all_media, tags, screenbloggers, fromdate, todate, dirname, platforms, followers_data, repformat,
                   brandname=brandname, campaign=campaign, reshares=reshares, blogposts=blogposts, post_info=post_info)
        excel(all_media, dirname, tags, followers_data)
        addtional_Excel(dirname)
    time.sleep(2)


if __name__ == '__main__':
    info = {'blogpost': {'Becca Rose': [], 'Chanel Boateng': [], 'Fashion Mumblr': ['http://www.fashionmumblr.com/2016/10/natural-way-finding-confidence-skin'], 'Hannah Maggs': []}, 'retweets': False, 'excludetags': ['Exclude tag'], 'comptags': [], 'competition': [], 'brandname': '',
            'replies': False, 'format': 'Full', 'blogposts': [], 'bloggers': ['Becca Rose', 'Chanel Boateng', 'Fashion Mumblr', 'Hannah Maggs'], 'todate': [2016, 11, 17], 'fromdate': [2016, 10, 19], 'tags': ['garnier'], 'campaign': '', 'platforms': ['Facebook', 'Instagram', 'Twitter', 'Blogs']}
  #  info= {'platforms': ['Facebook', 'Instagram', 'Twitter'], 'competition': [], 'replies': False, 'comptags': [], 'format': 'Full', 'fromdate': [2016, 6, 27], 'campaign': '', 'bloggers': ['Alpha Foodie', 'Amoureuse de Mode', 'Amy Valentine', 'Average Janes', 'Belle and Bunty', 'Chapters by S', 'Disneyrollergirl', 'Hannah Louise Fashion', "It's a LDN Thing", 'Julia Lundin', 'Karl Is My Unkle', 'LifenStyle', 'Lobler & Delaney', 'The Elgin Avenue', 'The Rollinson London'], 'todate': [2016, 7, 13], 'brandname': '', 'tags': ['Ceviche'], 'blogposts': [], 'blogpost': {}, 'excludetags': ['Exclude tag'], 'retweets': False}

  #  info= {'blogposts': [], 'fromdate': [2016, 2, 1], 'competition': [], 'tags': ['barbour', 'Latest on the blog: http://hannahlouisef.com/2016/05/white', 'New post! White Skinny Jeans', 'New post! - Black Waxed Jacket'], 'blogpost': {'The Elgin Avenue': ['http://www.theelginavenue.com/2016/02/low-key-lfw.html', 'http://www.theelginavenue.com/2016/05/sharing-a-failsafe-british-outfit-what-to-wear-in-the-rain-this-spring.html'], 'Framboise Fashion': ['http://www.framboisefashion.com/2016/02/girl-seeks-bike.html', 'http://www.framboisefashion.com/2016/05/cherry-armour.html'], 'Hannah Louise Fashion': ['http://hannahlouisef.com/2016/02/black-waxed-jacket/', 'http://hannahlouisef.com/2016/05/white-skinny-jeans/']}, 'brandname': '', 'retweets': False, 'platforms': ['Facebook', 'Instagram', 'Twitter'], 'campaign': '', 'format': 'Full', 'todate': [2016, 6, 7], 'excludetags': ['Exclude tag'], 'bloggers': ['Framboise Fashion', 'Hannah Louise Fashion', 'The Elgin Avenue'], 'comptags': [], 'replies': True}
  #  info= {'campaign': '', 'comptags': [], 'replies': False, 'todate': [2016, 6, 2], 'retweets': False, 'platforms': ['Instagram'], 'bloggers': ['Framboise Fashion', 'Hannah Louise Fashion', 'The Elgin Avenue'], 'blogposts': [], 'excludetags': ['Exclude tag'], 'fromdate': [2016, 2, 18], 'blogpost': {}, 'competition': [], 'tags': ['barbour'], 'brandname': '', 'format': 'Quick'}
  #  info= {'todate': [2016, 3, 15], 'fromdate': [2015, 12, 1], 'blogpost': {}, 'retweets': False, 'brandname': 'JD SPORTS', 'campaign': 'PINK SODA SPORT', 'competition': [], 'comptags': [], 'excludetags': ['Exclude tag'], 'blogposts': [], 'replies': False, 'bloggers': ['Fashion Influx'], 'format': 'Full', 'platforms': ['Instagram'], 'tags': ['a']}
    # info=pickle.load(open('{home}/Desktop/natappy/info.pickle'.format(home=home), 'rb'))
  #  info={'retweets': False, 'replies': False, 'blogpost': {'Julia Lundin': ['http://julialundin.com/la-redoute-carven-collection'], 'I Want You To Know': ['http://www.iwantyoutoknow.co.uk/2016/03/carven-la-redoute-designer-collaboration-review.html'], 'Camila Carril': ['http://www.camilacarril.com/2016/03/carven-la-redoute.html'], 'Stylonylon': ['http://stylonylon.com/2016/02/tous-les-garcons-et-les-filles.html'], 'Dreaming of Chanel': ['http://dreamingofchanel.com/new-designer-collaboration/', 'http://dreamingofchanel.com/the-pastel-jacket-ill-wear-again-and-again/']}, 'platforms': ['Facebook', 'Instagram', 'Twitter'], 'todate': [2016, 3, 11], 'brandname': 'La Redoute', 'comptags': [], 'fromdate': [2015, 11, 1], 'tags': ['#MyRedouteUK', 'laredoute_uk', 'laredouteuk'], 'blogposts': [], 'bloggers': ['Camila Carril', 'Dreaming of Chanel', 'I Want You To Know', 'Julia Lundin', 'Stylonylon'], 'format': 'Full', 'campaign': '', 'competition': [], 'excludetags': ['Exclude tag']}
  #  info= {'platforms': ['Facebook', 'Instagram', 'Twitter'], 'competition': [], 'excludetags': ['Exclude tag'], 'retweets': False, 'todate': [2016, 2, 17], 'bloggers': ['Hannah Louise Fashion'], 'blogpost': {}, 'comptags': [], 'tags': ['a'], 'format': 'Full', 'blogposts': [], 'replies': False, 'campaign': '', 'fromdate': [2016, 2, 14], 'brandname': ''}
    # info= {'tags': ['a'], 'format': 'Quick', 'campaign': '17 february', 'fromdate': [2016, 2, 10], 'blogposts': [], 'bloggers': ['Alice Point', 'Amelia Liana', 'Anneli Bush'], 'comptags': [], 'replies': False, 'blogpost': {}, 'retweets': False, 'excludetags': ['Exclude tag'], 'brandname': 'some report', 'platforms': ['Facebook', 'Instagram', 'Twitter'], 'competition': [], 'todate': [2016, 2, 17]}
    create_report(info['bloggers'], info['platforms'], info['tags'], info['format'], info['fromdate'], info['todate'], info[
                  'replies'], info['retweets'], info['excludetags'], info['brandname'], info['campaign'], info['blogpost'])
